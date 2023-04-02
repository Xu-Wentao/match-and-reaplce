import openpyxl

data_source_path = '中间库翻查_全媒介_313.xlsx'
template_source_path = '匹配模板.xlsx'

column_nums = []

tempaltes = []


def load_template():
	wb_obj = openpyxl.load_workbook(template_source_path)
	sheet_obj = wb_obj.active

	for n in sheet_obj.cell(row=2,column=1).value.split(' '):
		column_nums.append(int(n))

	row = 5
	while sheet_obj.cell(row=row,column=1).value:
		keys = sheet_obj.cell(row=row,column=1).value
		value_column = sheet_obj.cell(row=row,column=2).value
		value = sheet_obj.cell(row=row,column=3).value
		tempaltes.append({
			'keys':keys.split(' '),
			'value_column':int(value_column),
			'value':value
			})

		row +=1

def main():
	
	load_template()

	wb_obj = openpyxl.load_workbook(data_source_path)
	sheet_obj = wb_obj.active

	row = 3

	while sheet_obj.cell(row=row, column=1).value:
		print('第'+str(row)+'行')
		cells = [sheet_obj.cell(row=row, column=n).value for n in column_nums]

		for t in tempaltes:
			find = False

			for cell in cells:
				if not cell:
					continue
				for k in t['keys']:
					if k in cell:
						find = True
					if find:
						break
				if find:
					break

			if find:
				sheet_obj.cell(row, column=t['value_column']).value = t['value']

		row += 1

	wb_obj.save(''.join(data_source_path.split('.')[:-1])+str('_匹配导出.xlsx'))


if __name__ == '__main__':
	main()